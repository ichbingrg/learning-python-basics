import openpyxl as ox

inv_file = ox.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_inventory_price_per_supplier = {}
product_inventory_under_ten = {}

product_list.cell(1,5).value = "Total price"

for product_row in range(2, product_list.max_row+1):
    current_product_number = int(product_list.cell(product_row,1).value)
    current_inventory_number = int(product_list.cell(product_row,2).value)
    price = product_list.cell(product_row,3).value
    supplier_name = product_list.cell(product_row, 4).value

    #calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        #print(f"adding a new supplier : {supplier_name}")
        products_per_supplier[supplier_name] = 1

    #calculation for total price of inventory per supplier
    if supplier_name in total_inventory_price_per_supplier:
        total_inventory_price_per_supplier[supplier_name] += current_inventory_number*price
    else:
        #print(f"adding a new supplier for inventory: {supplier_name}")
        total_inventory_price_per_supplier[supplier_name] = current_inventory_number*price

    #printing out products with inventory less than ten
    if current_inventory_number<10:
        product_inventory_under_ten[current_product_number] = current_inventory_number

    #calculate the price of inventory per row and add it to the new column
    inventory_price = product_list.cell(product_row,5)
    inventory_price.value = current_inventory_number*price

print(f"products per supplier : {products_per_supplier}")
print(f"total inventory price per supplier : {total_inventory_price_per_supplier}")
print(f"products with inventory less than 10 : {product_inventory_under_ten}")

inv_file.save("new_inventory.xlsx")